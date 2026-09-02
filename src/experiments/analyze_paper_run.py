"""Read-only quantitative analysis for a live-paper strategy run.

The report deliberately separates signal quality (what the model knew) from
execution quality (what each 0/A-L policy did with the same signal).  It never
starts, stops, resets or otherwise changes the live paper runtime.
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from plotly import graph_objects as go
from plotly.io import to_html
from plotly.subplots import make_subplots


STARTING_CAPITAL = 100_000.0


def _timestamp(values: pd.Series) -> pd.Series:
    """State files contain ISO values both with and without microseconds."""
    return pd.to_datetime(values, utc=True, format="mixed", errors="coerce")


def _read_jsonl(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            # A concurrently-written final line may be incomplete.  The next run
            # will include it; keeping the report read-only is more important.
            continue
    return pd.DataFrame(rows)


def _run_states(run_dir: Path, run_id: str) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    score_rows: list[dict[str, Any]] = []
    trades_by_strategy: dict[str, pd.DataFrame] = {}
    history_by_strategy: dict[str, pd.DataFrame] = {}
    open_by_strategy: dict[str, pd.DataFrame] = {}
    for state_path in run_dir.glob("comparison_v1_*_*/state.json"):
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if state.get("run_id") != run_id:
            continue
        config, strategy_id = state.get("config", {}), state.get("config", {}).get("strategy_id")
        if not strategy_id:
            continue
        strategy_id = str(strategy_id)
        trades = pd.DataFrame(state.get("trades", []))
        if not trades.empty:
            trades = trades[trades.run_id.eq(run_id)].copy() if "run_id" in trades else trades.copy()
            trades["entry_datetime"] = _timestamp(trades.entry_time)
            trades["exit_datetime"] = _timestamp(trades.exit_time)
            trades["net_pnl"] = pd.to_numeric(trades.net_pnl, errors="coerce")
        history = pd.DataFrame(state.get("portfolio_history", []))
        if not history.empty:
            history = history[history.run_id.eq(run_id)].copy() if "run_id" in history else history.copy()
            history["datetime"] = _timestamp(history.timestamp)
            for field in ("balance", "equity", "realized_pnl", "unrealized_pnl", "max_drawdown", "exposure", "open_positions"):
                if field in history:
                    history[field] = pd.to_numeric(history[field], errors="coerce")
        trades_by_strategy[strategy_id] = trades
        history_by_strategy[strategy_id] = history
        positions = pd.DataFrame(state.get("positions", []))
        if not positions.empty:
            positions = positions[positions.run_id.eq(run_id)].copy() if "run_id" in positions else positions.copy()
            positions["entry_datetime"] = _timestamp(positions.entry_time)
            positions["exit_datetime"] = pd.NaT
            positions["exit_reason"] = "OPEN"
            positions["net_pnl"] = np.nan
        open_by_strategy[strategy_id] = positions
        realized, unrealized = float(state.get("realized_pnl", 0.0)), float(state.get("unrealized_pnl", 0.0))
        closed = trades.net_pnl.dropna() if not trades.empty else pd.Series(dtype=float)
        gains, losses = closed[closed > 0].sum(), abs(closed[closed < 0].sum())
        elapsed_hours = 0.0
        last_hour = np.nan
        if not history.empty:
            started, latest = history.datetime.min(), history.datetime.max()
            elapsed_hours = max((latest - started).total_seconds() / 3600, 0.0)
            before = history[history.datetime <= latest - pd.Timedelta(hours=1)]
            if not before.empty:
                last_hour = float(history.equity.iloc[-1] - before.equity.iloc[-1])
        score_rows.append({
            "strategy_id": strategy_id,
            "strategy": state.get("model", strategy_id),
            "total_pnl": realized + unrealized,
            "realized_pnl": realized,
            "unrealized_pnl": unrealized,
            "equity": float(state.get("equity", STARTING_CAPITAL)),
            "max_drawdown_pct": float(state.get("max_drawdown", 0.0)) * 100,
            "trades_closed": len(closed),
            "open_positions": len(state.get("positions", [])),
            "win_rate_pct": float((closed > 0).mean() * 100) if len(closed) else np.nan,
            "profit_factor": float(gains / losses) if losses else (np.inf if gains else 0.0),
            "expectancy": float(closed.mean()) if len(closed) else np.nan,
            "pnl_per_day": (realized + unrealized) / (elapsed_hours / 24) if elapsed_hours else np.nan,
            "pnl_last_hour": last_hour,
            "entry_mode": config.get("entry_mode"),
            "position_limit": config.get("max_open_positions"),
        })
    return pd.DataFrame(score_rows).sort_values("strategy_id"), trades_by_strategy, history_by_strategy, open_by_strategy


def _market_data(project_root: Path) -> pd.DataFrame:
    path = project_root / "data/live/MT5_XAUUSD_M1.parquet"
    market = pd.read_parquet(path)
    market["datetime_utc"] = _timestamp(market.datetime_utc)
    return market.sort_values("datetime_utc").drop_duplicates("datetime_utc", keep="last")


def _market_features(market: pd.DataFrame) -> pd.DataFrame:
    """Only backward-looking context available when the signal was created."""
    frame = market.copy().sort_values("datetime_utc")
    frame["mid_close"] = (frame.close_bid + frame.close_ask) / 2
    for minutes in (5, 15, 30, 60):
        frame[f"return_{minutes}m_pct"] = frame.mid_close.pct_change(minutes) * 100
    frame["range_5m"] = frame.mid_close.rolling(5).max() - frame.mid_close.rolling(5).min()
    frame["range_15m"] = frame.mid_close.rolling(15).max() - frame.mid_close.rolling(15).min()
    frame["hour_utc"] = frame.datetime_utc.dt.hour
    return frame


def _signal_quality(signals: pd.DataFrame, market: pd.DataFrame, minimum_move: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    if signals.empty:
        return pd.DataFrame(), pd.DataFrame()
    signals = signals.copy()
    signals["signal_datetime"] = _timestamp(signals.timestamp)
    signals["horizon_minutes"] = pd.to_numeric(signals.horizon_minutes, errors="coerce").fillna(15).astype(int)
    # Cost-aware live output keeps DOWN and NEUTRAL explicitly.  Its ``p_up``
    # is an execution score (0.4/0.5/0.6), therefore reconstruct the actual
    # UP probability rather than accidentally optimising that score.
    signals["p_up_raw"] = (1 - pd.to_numeric(signals.p_down, errors="coerce") - pd.to_numeric(signals.p_neutral, errors="coerce")).clip(0, 1)
    context = market[["datetime_utc", "return_5m_pct", "return_15m_pct", "return_30m_pct", "return_60m_pct", "range_5m", "range_15m", "hour_utc"]]
    signals = signals.merge(context, left_on="signal_datetime", right_on="datetime_utc", how="left").drop(columns="datetime_utc")
    # A signal is generated on a completed M1 and tradable at its captured BID/ASK.
    future = market[["datetime_utc", "close_bid", "close_ask"]].copy()
    future = future.rename(columns={"datetime_utc": "future_datetime", "close_bid": "future_bid", "close_ask": "future_ask"})
    signals["future_datetime"] = signals.signal_datetime + pd.to_timedelta(signals.horizon_minutes, unit="min")
    frame = signals.merge(future, on="future_datetime", how="left")
    frame["long_net_15m"] = frame.future_bid - pd.to_numeric(frame.ask, errors="coerce")
    frame["short_net_15m"] = pd.to_numeric(frame.bid, errors="coerce") - frame.future_ask
    frame["predicted_net_15m"] = np.select(
        [frame.predicted_class.eq("BUY"), frame.predicted_class.eq("SELL")],
        [frame.long_net_15m, frame.short_net_15m], default=0.0,
    )
    # Match the live model's target semantics: a BUY/SELL must clear a minimum
    # executable move after BID/ASK, otherwise the observed label is HOLD.
    frame["minimum_executable_move"] = minimum_move
    frame["actual_direction_15m"] = np.select(
        [frame.long_net_15m >= minimum_move, frame.short_net_15m >= minimum_move], ["BUY", "SELL"], default="HOLD"
    )
    frame["direction_correct"] = frame.predicted_class.eq(frame.actual_direction_15m)
    quality = frame.dropna(subset=["future_bid", "future_ask"]).copy()
    scorecard = quality.groupby("predicted_class", dropna=False).agg(
        signals=("signal_id", "count"),
        direction_accuracy_pct=("direction_correct", lambda x: float(x.mean() * 100)),
        tradable_win_rate_pct=("predicted_net_15m", lambda x: float((x > 0).mean() * 100)),
        total_theoretical_pnl=("predicted_net_15m", "sum"),
        average_theoretical_pnl=("predicted_net_15m", "mean"),
    ).reset_index()
    return quality, scorecard


def _probability_bins(signal_detail: pd.DataFrame, attribution: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    labels = ["<30%", "30-35%", "35-40%", "40-45%", "45-50%", "50-60%", "≥60%"]
    edges = [0, .30, .35, .40, .45, .50, .60, 1.000001]
    if signal_detail.empty:
        signal_bins = pd.DataFrame()
    else:
        signals = signal_detail[signal_detail.predicted_class.isin(["BUY", "SELL"])].copy()
        signals["directional_probability"] = np.where(signals.predicted_class.eq("BUY"), signals.p_up_raw, signals.p_down)
        signals["probability_band"] = pd.cut(signals.directional_probability, edges, labels=labels, right=False)
        signal_bins = signals.groupby(["predicted_class", "probability_band"], observed=False).agg(
            signals=("signal_id", "count"), theoretical_pnl_15m=("predicted_net_15m", "sum"),
            average_pnl_15m=("predicted_net_15m", "mean"), win_rate_pct=("predicted_net_15m", lambda x: float((x > 0).mean() * 100)),
        ).reset_index()
    if attribution.empty:
        return signal_bins, pd.DataFrame()
    trades = attribution.copy()
    trades["probability_band"] = pd.cut(trades.directional_probability, edges, labels=labels, right=False)
    trade_bins = trades.groupby(["strategy_id", "side", "probability_band"], observed=False).agg(
        trades=("trade_id", "count"), net_pnl=("net_pnl", "sum"), expectancy=("net_pnl", "mean"),
        win_rate_pct=("net_pnl", lambda x: float((x > 0).mean() * 100)),
    ).reset_index()
    return signal_bins, trade_bins


def _trade_attribution(trades_by_strategy: dict[str, pd.DataFrame], open_by_strategy: dict[str, pd.DataFrame], signals: pd.DataFrame, market: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    if signals.empty:
        return pd.DataFrame()
    signal_frame = signals.copy()
    signal_frame["signal_datetime"] = _timestamp(signal_frame.timestamp)
    signal_frame["p_up_raw"] = (1 - pd.to_numeric(signal_frame.p_down, errors="coerce") - pd.to_numeric(signal_frame.p_neutral, errors="coerce")).clip(0, 1)
    context = market[["datetime_utc", "return_5m_pct", "return_15m_pct", "return_30m_pct", "return_60m_pct", "range_5m", "range_15m", "hour_utc"]]
    signal_frame = signal_frame.merge(context, left_on="signal_datetime", right_on="datetime_utc", how="left").drop(columns="datetime_utc")
    keep = ["signal_id", "signal_datetime", "predicted_class", "p_down", "p_neutral", "p_up_raw", "spread",
            "return_5m_pct", "return_15m_pct", "return_30m_pct", "return_60m_pct", "range_5m", "range_15m", "hour_utc"]
    signal_frame = signal_frame[[column for column in keep if column in signal_frame]].drop_duplicates("signal_id", keep="last")
    for strategy_id, trades in trades_by_strategy.items():
        positions = open_by_strategy.get(strategy_id, pd.DataFrame())
        if trades.empty and positions.empty:
            continue
        frame = pd.concat([trades.copy(), positions.copy()], ignore_index=True, sort=False)
        frame["strategy_id"] = strategy_id
        frame["entry_datetime"] = _timestamp(frame["entry_datetime"])
        frame["exit_datetime"] = _timestamp(frame["exit_datetime"])
        frame["trade_status"] = np.where(frame.exit_datetime.notna(), "CLOSED", "OPEN")
        frame["holding_minutes"] = (frame.exit_datetime.fillna(pd.Timestamp.now(tz="UTC")) - frame.entry_datetime).dt.total_seconds().div(60).clip(lower=0)
        frame = frame.merge(signal_frame, on="signal_id", how="left")
        frame["directional_probability"] = np.where(frame.side.eq("LONG"), frame.p_up_raw, frame.p_down)
        frame["entry_hour_utc"] = frame.entry_datetime.dt.hour
        parts.append(frame)
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


def _run_candles(market: pd.DataFrame, signals: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if signals.empty:
        return pd.DataFrame(), pd.DataFrame()
    start = _timestamp(signals.timestamp).min().floor("min")
    end = _timestamp(signals.timestamp).max().ceil("min") + pd.Timedelta(minutes=1)
    m1 = market[(market.datetime_utc >= start) & (market.datetime_utc <= end)].copy()
    if m1.empty:
        return m1, pd.DataFrame()
    m5 = m1.set_index("datetime_utc").resample("5min").agg(
        open_bid=("open_bid", "first"), high_bid=("high_bid", "max"), low_bid=("low_bid", "min"), close_bid=("close_bid", "last"),
        open_ask=("open_ask", "first"), high_ask=("high_ask", "max"), low_ask=("low_ask", "min"), close_ask=("close_ask", "last"),
        tick_volume=("tick_volume", "sum"),
    ).dropna().reset_index()
    return m1, m5


def _write_trade_charts(output: Path, m1: pd.DataFrame, m5: pd.DataFrame, attribution: pd.DataFrame) -> None:
    """Write one zoomable HTML page per ledger, with M5 overview and M1 detail."""
    chart_folder = output / "trade_charts"
    chart_folder.mkdir(exist_ok=True)

    def figure(candles: pd.DataFrame, trades: pd.DataFrame, title: str) -> go.Figure:
        view = make_subplots(rows=2, cols=1, shared_xaxes=True, row_heights=[.78, .22], vertical_spacing=.03)
        view.add_trace(go.Candlestick(
            x=candles.datetime_utc, open=candles.open_bid, high=candles.high_bid, low=candles.low_bid, close=candles.close_bid,
            name="XAUUSD", increasing_line_color="#25b987", decreasing_line_color="#e45b6a",
        ), row=1, col=1)
        if "tick_volume" in candles:
            colours = np.where(candles.close_bid >= candles.open_bid, "#2f9e87", "#b95869")
            view.add_trace(go.Bar(x=candles.datetime_utc, y=candles.tick_volume, marker_color=colours, name="Tick volume", opacity=.55), row=2, col=1)
        if not trades.empty:
            for side, colour, symbol in (("LONG", "#25b987", "triangle-up"), ("SHORT", "#e45b6a", "triangle-down")):
                subset = trades[trades.side.eq(side)]
                if subset.empty:
                    continue
                text = [f"Trade #{row.trade_id}<br>{side} entry {row.entry_price:.2f}<br>PnL chiuso {row.net_pnl:+.2f} USD<br>P={row.directional_probability:.1%}" for row in subset.itertuples()]
                view.add_trace(go.Scatter(x=subset.entry_datetime, y=subset.entry_price, mode="markers", name=f"{side} entry",
                                          marker=dict(symbol=symbol, size=11, color=colour, line=dict(color="#ffffff", width=1)), text=text, hovertemplate="%{text}<extra></extra>"), row=1, col=1)
            closed = trades[trades.trade_status.eq("CLOSED")]
            exit_colours = np.where(closed.net_pnl >= 0, "#25b987", "#e45b6a")
            exit_text = [f"Trade #{row.trade_id}<br>EXIT {row.exit_price:.2f}<br>PnL {row.net_pnl:+.2f} USD<br>{row.exit_reason}" for row in closed.itertuples()]
            view.add_trace(go.Scatter(x=closed.exit_datetime, y=closed.exit_price, mode="markers", name="EXIT",
                                      marker=dict(symbol="x", size=11, color=exit_colours, line=dict(width=2)), text=exit_text, hovertemplate="%{text}<extra></extra>"), row=1, col=1)
        view.update_layout(template="plotly_dark", title=title, height=760, hovermode="closest", xaxis_rangeslider_visible=False,
                           legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0))
        view.update_yaxes(title_text="Prezzo", row=1, col=1)
        view.update_yaxes(title_text="Volume", row=2, col=1)
        return view

    links: list[str] = []
    for strategy_id, trades in attribution.groupby("strategy_id", sort=True) if not attribution.empty else []:
        title = f"Strategia {strategy_id} — M5 overview e M1 dettaglio"
        page = "<h1 style='font-family:Arial'>" + title + "</h1>"
        page += "<p style='font-family:Arial'>Zoom: rotella / selezione. Hover su marker per trade, PnL, probabilità e motivo di uscita.</p>"
        page += "<h2 style='font-family:Arial'>M5 — panoramica</h2>" + to_html(figure(m5, trades, title + " (M5)"), include_plotlyjs="cdn", full_html=False)
        page += "<h2 style='font-family:Arial'>M1 — dettaglio</h2>" + to_html(figure(m1, trades, title + " (M1)"), include_plotlyjs=False, full_html=False)
        filename = f"strategy_{strategy_id}.html"
        (chart_folder / filename).write_text("<html><body style='background:#0e1626;color:#edf2f7;padding:20px'>" + page + "</body></html>", encoding="utf-8")
        links.append(f"<li><a href='{filename}' style='color:#64d8ff'>Strategia {strategy_id}</a></li>")
    (chart_folder / "index.html").write_text(
        "<html><body style='background:#0e1626;color:#edf2f7;font-family:Arial;padding:30px'><h1>Trade charts</h1><p>Grafici interattivi generati da uno snapshot read-only.</p><ul>" + "".join(links) + "</ul></body></html>",
        encoding="utf-8",
    )


def _long_short_hourly(trades_by_strategy: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame]:
    parts: list[pd.DataFrame] = []
    for strategy_id, trades in trades_by_strategy.items():
        if trades.empty:
            continue
        frame = trades.copy()
        frame["strategy_id"] = strategy_id
        frame["exit_hour_utc"] = frame.exit_datetime.dt.hour
        parts.append(frame)
    if not parts:
        return pd.DataFrame(), pd.DataFrame()
    all_trades = pd.concat(parts, ignore_index=True)
    long_short = all_trades.groupby(["strategy_id", "side"], dropna=False).agg(
        trades=("trade_id", "count"), net_pnl=("net_pnl", "sum"), expectancy=("net_pnl", "mean"),
        win_rate_pct=("net_pnl", lambda x: float((x > 0).mean() * 100)),
    ).reset_index()
    hourly = all_trades.groupby(["strategy_id", "exit_hour_utc"], dropna=False).agg(
        trades=("trade_id", "count"), net_pnl=("net_pnl", "sum"), expectancy=("net_pnl", "mean"),
        win_rate_pct=("net_pnl", lambda x: float((x > 0).mean() * 100)),
    ).reset_index()
    return long_short, hourly


def _reversal_counterfactual(trades_by_strategy: dict[str, pd.DataFrame], market: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    details: list[dict[str, Any]] = []
    for strategy_id, trades in trades_by_strategy.items():
        if trades.empty or "exit_reason" not in trades:
            continue
        candidate = trades[(trades.side.eq("SHORT")) & (trades.exit_reason.eq("smart short confirmed reversal"))]
        for trade in candidate.itertuples(index=False):
            start = trade.exit_datetime.floor("min") + pd.Timedelta(minutes=1)
            future = market[market.datetime_utc >= start]
            outcome, hit_time = "unresolved", pd.NaT
            for bar in future.itertuples(index=False):
                hit_tp = float(bar.low_ask) <= float(trade.take_profit)
                hit_sl = float(bar.high_ask) >= float(trade.stop_loss)
                if hit_tp or hit_sl:
                    # M1 cannot reveal the order when both are touched.  Stop-first is conservative.
                    outcome = "tp_first" if hit_tp and not hit_sl else "sl_first"
                    hit_time = bar.datetime_utc
                    break
            hypothetical = np.nan
            if outcome == "tp_first":
                hypothetical = float(trade.raw_entry_price) - float(trade.take_profit) - float(trade.costs)
            elif outcome == "sl_first":
                hypothetical = float(trade.raw_entry_price) - float(trade.stop_loss) - float(trade.costs)
            details.append({
                "strategy_id": strategy_id, "trade_id": trade.trade_id, "exit_time": trade.exit_time,
                "actual_net_pnl": float(trade.net_pnl), "outcome_if_held": outcome, "hit_time": hit_time,
                "hypothetical_net_pnl": hypothetical,
                "incremental_pnl_if_held": hypothetical - float(trade.net_pnl) if pd.notna(hypothetical) else np.nan,
            })
    detail = pd.DataFrame(details)
    if detail.empty:
        return detail, pd.DataFrame()
    summary = detail.groupby(["strategy_id", "outcome_if_held"], dropna=False).agg(
        trades=("trade_id", "count"), actual_net_pnl=("actual_net_pnl", "sum"),
        hypothetical_net_pnl=("hypothetical_net_pnl", "sum"), incremental_pnl_if_held=("incremental_pnl_if_held", "sum"),
    ).reset_index()
    return detail, summary


def _shock_impact(market: pd.DataFrame, history_by_strategy: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame]:
    market = market.copy()
    market["m1_change"] = market.close_bid.diff()
    active = market.dropna(subset=["m1_change"])
    threshold = max(2.5, float(active.m1_change.abs().quantile(.99))) if not active.empty else np.inf
    shocks = active[active.m1_change.abs() >= threshold].copy()
    shocks["shock_threshold"] = threshold
    impact: list[dict[str, Any]] = []
    for shock in shocks.itertuples(index=False):
        for strategy_id, history in history_by_strategy.items():
            if history.empty:
                continue
            before = history[history.datetime <= shock.datetime_utc - pd.Timedelta(minutes=1)]
            after = history[history.datetime <= shock.datetime_utc + pd.Timedelta(minutes=15)]
            if before.empty or after.empty:
                continue
            impact.append({
                "strategy_id": strategy_id, "shock_time": shock.datetime_utc, "m1_change": shock.m1_change,
                "equity_before": float(before.equity.iloc[-1]), "equity_after_15m": float(after.equity.iloc[-1]),
                "pnl_impact_15m": float(after.equity.iloc[-1] - before.equity.iloc[-1]),
            })
    return shocks, pd.DataFrame(impact)


def _correlations(history_by_strategy: dict[str, pd.DataFrame]) -> pd.DataFrame:
    series: list[pd.DataFrame] = []
    for strategy_id, history in history_by_strategy.items():
        if history.empty:
            continue
        item = history[["datetime", "equity"]].dropna().drop_duplicates("datetime", keep="last").copy()
        item["minute"] = item.datetime.dt.floor("min")
        item = item.groupby("minute", as_index=False).equity.last().set_index("minute").pct_change()
        series.append(item.rename(columns={"equity": strategy_id}))
    if not series:
        return pd.DataFrame()
    aligned = pd.concat(series, axis=1, sort=False).dropna(how="all")
    return aligned.corr().reset_index(names="strategy_id")


def _write_excel(output: Path, tables: dict[str, pd.DataFrame], metadata: dict[str, str]) -> None:
    workbook_path = output / "paper_analysis.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(list(metadata.items()), columns=["Campo", "Valore"]).to_excel(writer, sheet_name="README", index=False)
        for name, frame in tables.items():
            if frame is None or frame.empty:
                continue
            excel_frame = frame.copy()
            # Excel has no timezone representation.  CSV keeps the original UTC
            # ISO values; the workbook uses timezone-naive UTC timestamps.
            for column in excel_frame.columns:
                if isinstance(excel_frame[column].dtype, pd.DatetimeTZDtype):
                    excel_frame[column] = excel_frame[column].dt.tz_convert("UTC").dt.tz_localize(None)
            excel_frame.to_excel(writer, sheet_name=name[:31], index=False)
    book = load_workbook(workbook_path)
    navy, blue, green = "12213A", "1F4E78", "D9EAD3"
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells[:500]) + 2, 28)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        if sheet.max_row > 1 and sheet.max_column > 1:
            table = Table(displayName=f"tbl_{sheet.title.replace(' ', '_').replace('-', '_')}", ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
    if "Strategy scorecard" in book.sheetnames:
        sheet = book["Strategy scorecard"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        pnl_col = headers.get("total_pnl")
        id_col = headers.get("strategy_id")
        if pnl_col and id_col:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title, chart.y_axis.title, chart.x_axis.title = "PnL totale per strategia", "USD", "Strategia"
            chart.add_data(Reference(sheet, min_col=pnl_col, min_row=1, max_row=sheet.max_row), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=id_col, min_row=2, max_row=sheet.max_row))
            chart.height, chart.width = 8, 16
            chart.legend = None
            # A single Excel series otherwise defaults to black/white.  Colour each
            # strategy by sign so the chart is readable at a glance.
            values = [sheet.cell(row=row, column=pnl_col).value for row in range(2, sheet.max_row + 1)]
            for index, value in enumerate(values):
                point = DataPoint(idx=index)
                point.graphicalProperties.solidFill = "25B987" if float(value or 0) >= 0 else "E45B6A"
                point.graphicalProperties.line.solidFill = "25B987" if float(value or 0) >= 0 else "E45B6A"
                chart.series[0].dPt.append(point)
            sheet.add_chart(chart, "S2")
            pnl_letter = get_column_letter(pnl_col)
            sheet.conditional_formatting.add(f"{pnl_letter}2:{pnl_letter}{sheet.max_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    if "Hourly chart" in book.sheetnames:
        sheet = book["Hourly chart"]
        if sheet.max_column > 1:
            chart = LineChart()
            chart.style = 13
            chart.title, chart.y_axis.title, chart.x_axis.title = "PnL realizzato per ora UTC", "USD", "Ora"
            chart.add_data(Reference(sheet, min_col=2, max_col=sheet.max_column, min_row=1, max_row=sheet.max_row), titles_from_data=True, from_rows=False)
            chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row))
            chart.height, chart.width = 8, 16
            sheet.add_chart(chart, "A28")
    book.save(workbook_path)


def analyze(project_root: Path, run_dir: Path | None = None, output_root: Path | None = None, minimum_move: float = .5) -> Path:
    run_dir = run_dir or project_root / "data/live/paper/comparison_v1"
    metadata = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
    run_id = str(metadata["run_id"])
    # A raw ISO timestamp is useful to software but tells a human nothing in
    # Explorer.  Keep snapshots separate, but give every new folder a readable
    # name explaining both its content and when it was generated.
    generated_at = datetime.now(timezone.utc)
    analysis_root = project_root / "results/paper_analysis" / run_id
    friendly_snapshot = generated_at.strftime("ANALISI_PAPER_%Y-%m-%d__%H-%M-%S_UTC")
    output = output_root or analysis_root / friendly_snapshot
    output.mkdir(parents=True, exist_ok=True)
    scorecard, trades, history, open_positions = _run_states(run_dir, run_id)
    market, signals = _market_features(_market_data(project_root)), _read_jsonl(run_dir / "signals.jsonl")
    if not signals.empty and "run_id" in signals:
        signals = signals[signals.run_id.eq(run_id)].copy()
    signal_detail, signal_scorecard = _signal_quality(signals, market, minimum_move)
    attribution = _trade_attribution(trades, open_positions, signals, market)
    signal_bins, trade_bins = _probability_bins(signal_detail, attribution)
    candles_m1, candles_m5 = _run_candles(market, signals)
    long_short, hourly = _long_short_hourly(trades)
    hourly_chart = (hourly.pivot(index="exit_hour_utc", columns="strategy_id", values="net_pnl").fillna(0).reset_index()
                    if not hourly.empty else pd.DataFrame())
    reversal_detail, reversal_summary = _reversal_counterfactual(trades, market)
    shocks, shock_impact = _shock_impact(market, history)
    correlation = _correlations(history)
    tables = {
        "Strategy scorecard": scorecard, "Model quality 15m": signal_scorecard,
        "Signal detail 15m": signal_detail, "Long short": long_short, "Hourly": hourly, "Hourly chart": hourly_chart,
        "Probability signal bins": signal_bins, "Probability trade bins": trade_bins,
        "Trade attribution": attribution, "Candles M1": candles_m1, "Candles M5": candles_m5,
        "Reversal summary": reversal_summary, "Reversal detail": reversal_detail,
        "Shock windows": shocks, "Shock impact": shock_impact, "Correlation": correlation,
    }
    for title, frame in tables.items():
        if frame is not None and not frame.empty:
            frame.to_csv(output / f"{title.lower().replace(' ', '_')}.csv", index=False)
    metadata_out = {
        "run_id": run_id, "generated_at_utc": generated_at.isoformat(),
        "analysis_mode": "read-only snapshot; live runtime unchanged",
        "market_m1_last_bar": str(market.datetime_utc.max()),
        "signals_analysed": str(len(signal_detail)), "strategies_analysed": str(len(scorecard)),
        "signal_quality_minimum_executable_move": str(minimum_move),
        "probability_note": "p_up_raw = 1 - p_down - p_neutral; the logged p_up field is only the operational 0.4/0.5/0.6 score.",
        "reversal_note": "TP/SL order within one M1 is treated conservatively as stop-first.",
    }
    (output / "analysis_metadata.json").write_text(json.dumps(metadata_out, indent=2), encoding="utf-8")
    if scorecard.empty:
        summary = "No strategies found."
    else:
        columns = ["strategy_id", "total_pnl", "profit_factor", "max_drawdown_pct", "trades_closed", "pnl_last_hour"]
        # ``DataFrame.to_markdown`` needs the optional tabulate package.  A CSV
        # code block is portable and keeps this reporting command dependency-light.
        summary = "```csv\n" + scorecard[columns].to_csv(index=False) + "```"
    (output / "decision_summary.md").write_text(
        "# Paper analysis\n\nThis report is descriptive, not a promotion to live trading.\n\n" + summary + "\n",
        encoding="utf-8",
    )
    _write_excel(output, tables, metadata_out)
    _write_trade_charts(output, candles_m1, candles_m5, attribution)
    if output_root is None:
        (analysis_root / "LEGGIMI - QUALE CARTELLA APRIRE.txt").write_text(
            "Questa cartella contiene piu snapshot della STESSA sessione Paper.\n\n"
            "Apri sempre la cartella con il nome piu recente che inizia con:\n"
            "  ANALISI_PAPER_YYYY-MM-DD__HH-MM-SS_UTC\n\n"
            f"L'ultima analisi creata e:\n  {output.name}\n\n"
            "Dentro trovi prima di tutto:\n"
            "  - paper_analysis.xlsx  (riepilogo leggibile)\n"
            "  - trade_charts\\index.html  (grafici interattivi)\n"
            "Le vecchie cartelle che iniziano solo con numeri sono snapshot storici\n"
            "automatici dello stesso run: non sono strategie o esperimenti diversi.\n",
            encoding="utf-8",
        )
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description="Read-only analysis and Excel report for the current paper run")
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    parser.add_argument("--run-dir", type=Path)
    parser.add_argument("--output-root", type=Path)
    parser.add_argument("--minimum-move", type=float, default=.5,
                        help="executable 15m move used to label BUY/SELL; default matches active H15 move0.5 model")
    args = parser.parse_args()
    result = analyze(args.project_root.resolve(), args.run_dir.resolve() if args.run_dir else None, args.output_root.resolve() if args.output_root else None, args.minimum_move)
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
